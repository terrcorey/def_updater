import xlwings as xw
import sys
import os
import numpy as np
import pandas as pd
import json
from tqdm import tqdm
import openpyxl
from openpyxl.utils import get_column_letter

def format_excel(output_excel_path):
    print("Formatting the Excel file...")
    try:
        # Open workbook with openpyxl for autofit
        wb = openpyxl.load_workbook(output_excel_path)
        for ws in wb.worksheets:
            for idx in range(1, ws.max_column + 1):
                col_letter = get_column_letter(idx)
                ws.column_dimensions[col_letter].width = 12
        wb.save(output_excel_path)
        wb.close()
    except Exception as e:
        print(f"Error formatting workbook: {e}")

def create_excel_file():
    excel_file_path = os.path.join(".", "label_editor.xlsx")
    # Check if the Excel file exists
    if os.path.exists(excel_file_path):
        confirm = input(f"{excel_file_path} already exists. Are you sure you want to regenerate a new file? (y/n): ").strip().lower()
        if confirm != 'y':
            print("Operation cancelled.")
            sys.exit(0)
        os.remove(excel_file_path)


def read_states(states_file_path):
    # Read the states file into a DataFrame
    df = pd.read_csv(states_file_path, sep=r"\s+", header=None, nrows=1)
    return df

def write_to_excel(df, desc_df, mol, filename):
    ds_name = filename.replace(".def", "")
    excel_file_path = os.path.join(".", "label_editor.xlsx")
    sheet_name = f"{mol}__{ds_name}"
    sheet_name = sheet_name[:31]
    try:
        with pd.ExcelWriter(excel_file_path, mode='a' if os.path.exists(excel_file_path) else 'w', engine='openpyxl') as writer:
            df.to_excel(writer, sheet_name=sheet_name, index=False, header=False)
            ws = writer.sheets[sheet_name]
            ws.insert_rows(1)
            ws.cell(row=1, column=1, value=f"Molecule: {mol}, File: {filename}")
            # Write desc_df to the sixth row (row 6, columns starting from 1)
            for col_idx, value in enumerate(desc_df.iloc[0, :], start=1):
                ws.cell(row=col_idx + 6, column=1, value=value)
            for col_idx, value in enumerate(desc_df.iloc[1, :], start=1):
                ws.cell(row=col_idx + 6, column=2, value=value)
    except Exception as e:
        print(f"{filename} || Error saving DataFrame to Excel: {e}")

def read_def_labels(path):
    with open(path, "r") as f:
        lines = f.readlines()
    quantum_labels = [
        {
            "Quantum label": "ID",
            "Description quantum label": "Unique integer identifier for the energy level",
            "Format quantum label": "I12 %12d",
        },
        {
            "Quantum label": "E",
            "Description quantum label": "State energy in cm-1",
            "Format quantum label": "F12.6 %12.6f",
        },
        {
            "Quantum label": "gtot",
            "Description quantum label": "Total energy level degeneracy",
            "Format quantum label": "I6 %6d"
        },
        {
            "Quantum label": "J",
            "Description quantum label": "Total rotational quantum number, excluding nuclear spin",
            "Format quantum label": "I7 %7d"
        }
    ]
    lines_iter = iter(lines)
    for l in lines_iter:
        if "# Quantum label" in l:
            quantum_label = l.split("#")[0].strip()
            fmt_label = next(lines_iter).split("#")[0].strip()
            desc_label = next(lines_iter).split("#")[0].strip()
            quantum_labels.append({
                "Quantum label": quantum_label,
                "Format quantum label": fmt_label,
                "Description quantum label": desc_label
            })
        if "# Lifetime availability (1=yes, 0=no)" in l:
            lifetime_avail = l.split("#")[0].strip()
            gfactor_avail = next(lines_iter).split("#")[0].strip()
            unc_avail = next(lines_iter).split("#")[0].strip()
            hyperfine_dataset = next(lines_iter).split("#")[0].strip()
            insert_idx = 4
            if unc_avail == "1":
                quantum_labels.insert(insert_idx, {
                    "Quantum label": "unc",
                    "Format quantum label": "F12.6 %12.6f",
                    "Description quantum label": "Energy uncertainty in cm-1"
                })
                insert_idx += 1
            if lifetime_avail == "1":
                quantum_labels.insert(insert_idx, {
                    "Quantum label": "tau",
                    "Format quantum label": "ES12.4 %12.4e",
                    "Description quantum label": "Lifetime in s"
                })
                insert_idx += 1
            if gfactor_avail == "1":
                quantum_labels.insert(insert_idx, {
                    "Quantum label": "gfactor",
                    "Format quantum label": "F10.6 %10.6f",
                    "Description quantum label": "Landé g-factor"
                })
            if hyperfine_dataset == "1":
                quantum_labels[3] = {
                    "Quantum label": "F",
                    "Description quantum label": "Final angular momentum quantum number",
                    "Format quantum label": "I7 %7d"
                }
        if "# Auxiliary title" in l:
            auxiliary_title = l.split("#")[0].strip()
            fmt_label = next(lines_iter).split("#")[0].strip()
            desc_label = next(lines_iter).split("#")[0].strip()
            quantum_labels.append({
                "Quantum label": "Auxiliary:" + auxiliary_title,
                "Format quantum label": fmt_label,
                "Description quantum label": desc_label
            })
    return quantum_labels

def read_json_labels(path):
    with open(path, "r") as f:
        data = json.load(f)
    quanta = data['dataset']['states'].get("states_file_fields", [])
    quantum_labels = []
    for q in quanta:
        quantum_labels.append({
            "Quantum label": q.get("name", ""),
            "Format quantum label": " ".join([q.get("ffmt", ""), q.get("cfmt", "")]),
            "Description quantum label": q.get("desc", "")
        })
    return quantum_labels

def main():
    input_path = os.path.join(".", "input")
    mol_dirs = [d for d in os.listdir(input_path) if os.path.isdir(os.path.join(input_path, d))]
    for mol in tqdm(mol_dirs, desc = "Printing labels onto Excel..."):
        # Process each molecule directory
        mol_path = os.path.join(input_path, mol)
        for f in os.listdir(mol_path):
            def_file_path = os.path.join(mol_path, f)
            if def_file_path.endswith(".def"):
                states_file_path = os.path.join(mol_path, f.replace(".def", ".states"))
                if os.path.exists(states_file_path):
                    states_df = read_states(states_file_path)
                else:
                    print(f"States file not found for {def_file_path}. Skipping.")
                    continue
                # Read the labels from the definition file
                #if os.path.exists(def_file_path.replace(".def", ".def.json")):
                    #def_labels = read_json_labels(def_file_path.replace(".def", ".def.json"))
                #else:
                def_labels = read_def_labels(def_file_path)
                if not isinstance(states_df.iloc[0, 3], int):
                    def_labels[3]["Format quantum label"] = "F7.1 %7.1f"

                # Stack def_labels and states_df
                labels = [l['Quantum label'] for l in def_labels]
                fmt = [l['Format quantum label'] for l in def_labels]
                description = [l['Description quantum label'] for l in def_labels]
                header_df = pd.DataFrame([labels, fmt])
                full_df = pd.concat([header_df, states_df], ignore_index=True)
                desc_df = pd.DataFrame([labels, description])
                full_df = full_df.replace({np.nan: None})
                write_to_excel(full_df, desc_df, mol, f)
    

if __name__ == "__main__":
    # Define the path to the Excel file
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    create_excel_file()
    main()
    format_excel(os.path.join(".", "label_editor.xlsx"))
