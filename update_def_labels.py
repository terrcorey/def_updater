import numpy as np
import os
import sys
from tqdm import tqdm
import json
from datetime import datetime
import inspect
import pandas as pd
import re
import requests
import bz2
from io import BytesIO
import openpyxl
from pprint import pprint

def download_states_first_line(mol, f):
    if f.endswith(".def"):
        f.replace(".def", "")
    states_dir = os.path.join(".", "input", mol)
    states_file_path = os.path.join(states_dir, f + ".states")
    if not os.path.exists(states_file_path):
        iso_slug, ds_name = f.split("__")
        url = f"https://exomol.com/db/{mol}/{iso_slug}/{ds_name}/{f + '.states.bz2'}".replace("+", "_p")
        response = requests.get(url)
        if response.status_code == 200:
            # Decompress only until the first 10 lines are read
            decompressor = bz2.BZ2Decompressor()
            buffer = BytesIO(response.content)
            lines = []
            line_buffer = b""
            while len(lines) < 10:
                chunk = buffer.read(1024)
                if not chunk:
                    break
                data = decompressor.decompress(chunk)
                line_buffer += data
            while b"\n" in line_buffer and len(lines) < 10:
                line, line_buffer = line_buffer.split(b"\n", 1)
                lines.append(line)
            # Save the first 10 lines to the states file
            with open(states_file_path, "wb") as f:
                for line in lines:
                    f.write(line + b"\n")

        else:
            error_log(f"Dataset: {f.replace('.def', '')} || Failed to download states file from {url}", "Error")

def check_J_format(states_file_path):
    # Reads only the first row to check the J column format
    df = pd.read_csv(states_file_path, sep=r"\s+", header=None, nrows=1)
    # Assume J is in the 4th column (index 3)
    j_col = df.iloc[:, 3]
    # If any value is not integer, use float format
    if any(j_col.apply(lambda x: not float(x).is_integer())):
        return "%7.1f", "F7.1"
    return "%7d", "I7"

def generate_formats(number_with_spaces):
    """
    Generates Fortran and C formats for a given number string with leading spaces,
    including support for scientific notation.
    
    Args:
        number_with_spaces (str): The number as a string, including leading spaces.
    
    Returns:
        tuple: A tuple containing the Fortran format and C format strings.
    """
    number = number_with_spaces.strip()
    leading_spaces = len(number_with_spaces) - len(number) - 1
    
    try:
        # Check if the number is an integer
        value = int(number)
        val_len = len(str(value))
        buff = val_len + leading_spaces
        fortran_format = f"I{buff}"
        c_format = f"%{buff}d"
        return fortran_format, c_format
    except ValueError:
        try:
            # Check if the number is a float
            value = float(number)
            val_len_int, val_len_frac = [len(x) for x in str(value).split(".")]
            buff = val_len_int + leading_spaces
            fortran_format = f"F{buff}.{val_len_frac}"
            c_format = f"%{buff}.{val_len_frac}f"
            return fortran_format, c_format
        except ValueError:
            try:
                # Check if the number is in scientific notation
                value = float(number)
                val_len = len(number.replace("e", "").replace("+", "").replace("-", ""))
                buff = val_len + leading_spaces
                fortran_format = f"E{buff}.{val_len - 1}"
                c_format = f"%{buff}.{val_len - 1}e"
                return fortran_format, c_format
            except ValueError:
                # If the number is a string
                val_len = len(number)
                buff = val_len + leading_spaces
                fortran_format = f"A{buff}"
                c_format = f"%{buff}s"
                return fortran_format, c_format

def detect_format(states_file_path, idx):
    """Detects the format of the quantum label based on the states file."""
    with open(states_file_path, 'r') as f:
        first_line = f.readline().rstrip('\n')
    # Find all columns and their leading spaces
    matches = list(re.finditer(r'( +)?(\S+)', first_line))
    idx = idx + 4
    if idx < len(matches):
        match = matches[idx]
        match_str = match.group(1) + match.group(2)
        ffmt, cfmt = generate_formats(match_str)
        return ffmt + " " + cfmt
    else:
        error_log(f"Dataset: {filename} || Quantum label index {idx} exceeds the number of columns in the states file.", "Critical")
        return ""
    
# json loaders
def load_correction_dict():
    path = os.path.join(os.getcwd(), "other_materials", "lib", "correction_dict.json")
    with open(path, "r", encoding="utf-8") as f:
        return json.load(f)
    
def load_labels_data():
    path = os.path.join(os.getcwd(), "other_materials", "lib", "labels.json")
    #path = os.path.join(os.getcwd(), "states_labels.json")
    with open(path, "r", encoding="utf-8") as f:
        labels_data = json.load(f)
    return labels_data

def load_standard_labels():
    path = os.path.join(os.getcwd(), "other_materials", "lib", "standard_label_structure.json")
    with open(path, "r", encoding="utf-8") as f:
        standard_labels = json.load(f)
    return standard_labels

def load_backend_master_file():
    path = os.path.join(os.getcwd(), "other_materials", "lib", "exomol-20250905.all.json")
    with open(path, "r", encoding="utf-8") as f:
        old_master_file = json.load(f)
    return old_master_file

def make_def_json():
    """Creates a JSON file with the standard labels structure."""
    path = os.path.join(os.getcwd(), "other_materials", "scripts", "convert_newnew.py")
    os.system(f"python {path}")

# Error handling and logging
def exit_script():
    log_file_path = os.path.join(os.getcwd(), "log.txt")
    with open(log_file_path, 'r') as log_file:
        log_content = log_file.readlines()
    w_counter = 0
    e_counter = 0
    c_counter = 0
    for l in log_content:
        if "[Warn]" in l:
            w_counter += 1
        elif "[Error]" in l:
            e_counter += 1
        elif "[Critical]" in l:
            c_counter += 1
    print(f"Script execution completed with {w_counter} warnings, {e_counter} errors, and {c_counter} critical errors.")
    print("Please check the log file for details.")
    print("Exiting script...")
    error_log("Script ended at " + datetime.now().strftime("%Y-%m-%d %H:%M:%S"), "Log")
    sys.exit(0)

def error_log(message, prefix="Error"):
    """Logs an error message to a file, including the full call stack line numbers (formatted for readability)."""
    log_file_path = os.path.join(os.getcwd(), "log.txt")
    os.makedirs(os.path.dirname(log_file_path), exist_ok=True)
    # Build call stack info, each frame on a new line, indented
    if prefix == "Log":
        with open(log_file_path, 'a') as log_file:
            log_file.write(f"[{prefix}] at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} {message}\n")
    else:
        stack = inspect.stack()
        stack_lines = []
        for frame in reversed(stack[1:]):
            stack_lines.append(f"    {frame.function}@{frame.lineno}")
        stack_info = "\n" + "\n".join(stack_lines)
        with open(log_file_path, 'a') as log_file:
            log_file.write(f"[{prefix}] at {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}{stack_info}\n    {message}\n")

# Input reading and processing
def read_def_file(def_file_path):
    """Reads a definition file and returns a dictionary mapping descriptions to values."""
    correction_dict = load_correction_dict()
    def_dict = {}

    with open(def_file_path, 'r') as def_file:
        lines = def_file.readlines()

    for l in lines:
        if "#" not in l:
            continue
        parts = [entry.strip() for entry in l.split("#", 1)]
        if len(parts) != 2:
            continue
        value, desc = parts
        if ": bound/quasi-bound" in desc:
            desc = desc.replace(": bound/quasi-bound", "")
        try:
            _ = int(desc[-2:])
            desc = desc[:-2].strip()
        except (ValueError, IndexError):
            pass
        if value in correction_dict:
            value = correction_dict[value]
        if def_dict.get(desc):
            temp = def_dict[desc]
            if isinstance(temp, list):
                temp.append(value)
            else:
                def_dict[desc] = [temp, value]
        else:
            def_dict[desc] = value

    filename = os.path.basename(def_file_path).replace(".def", "").split("__")
    def_dict["Iso-slug"] = filename[0]
    def_dict["Isotopologue dataset name"] = filename[1]

    quantum_labels = []
    auxiliary_labels = []

    label = def_dict.get("Quantum label", [])
    fmt = def_dict.get("Format quantum label", [])
    desc = def_dict.get("Description quantum label", [])
    if isinstance(label, str):
        label = [label]
        fmt = [fmt]
        desc = [desc]

    for i, l in enumerate(label):
        quantum_labels.append({
            "Quantum label": l,
            "Format quantum label": fmt[i],
            "Description quantum label": desc[i]
        })
    def_dict.pop("Quantum label", None)
    def_dict.pop("Format quantum label", None)
    def_dict.pop("Description quantum label", None)

    aux = def_dict.get("Auxiliary title", [])
    if aux == []:
        aux = def_dict.get("Auxiliary tite", [])
    fmt_aux = def_dict.get("Format title", [])
    desc_aux = def_dict.get("Description title", [])
    if isinstance(aux, str):
        aux = [aux]
        fmt_aux = [fmt_aux]
        desc_aux = [desc_aux]

    for i, a in enumerate(aux):
        auxiliary_labels.append({
            "Auxiliary title": a,
            "Format title": fmt_aux[i],
            "Description title": desc_aux[i]
        })
    def_dict.pop("Auxiliary title", None)
    def_dict.pop("Auxiliary tite", None)
    def_dict.pop("Format title", None)
    def_dict.pop("Description title", None)

    isotopes = def_dict.get("Isotope number", [])
    symbol = def_dict.get("Element symbol", [])
    if isinstance(isotopes, str):
        isotopes = [isotopes]
        symbol = [symbol]

    isotope_info = []
    for i, l in enumerate(isotopes):
        isotope_info.append({
            "Isotope number": l,
            "Element symbol": symbol[i] 
        })
    def_dict["Isotope information"] = isotope_info
    def_dict.pop("Isotope number", None)
    def_dict.pop("Element symbol", None)

    irreps_list = def_dict["Irreducible representation label"]
    nuc_deg_list = def_dict["Nuclear spin degeneracy"]
    try:
        idx_list = def_dict["Irreducible representation ID"] 
        assert len(idx_list) == len(irreps_list) == len(nuc_deg_list)
    except KeyError:
        pass
    assert len(irreps_list) == len(nuc_deg_list), "Irreducible representation lists are not of the same length."
    irreps = []
    for i, l in enumerate(irreps_list):
        irreps.append({
            "Irreducible representation label": l,
            "Nuclear spin degeneracy": nuc_deg_list[i]
        })
    def_dict["irreps"] = irreps
    def_dict.pop("Irreducible representation ID", None)
    def_dict.pop("Irreducible representation label", None)
    def_dict.pop("Nuclear spin degeneracy", None)

    label_list = def_dict.get("Label for a particular broadener", None)
    fname_list = def_dict.get("Filename of particular broadener", None)
    if fname_list == None:
        fname_list = def_dict.get("Filename for a particular broadener", None)
    max_J_list = def_dict.get("Maximum J for which pressure broadening parameters provided", None)
    lor_hw_list = def_dict.get('Value of Lorentzian half-width for J" > Jmax', None)
    temp_exp_list = def_dict.get('Value of temperature exponent for lines with J" > Jmax', None)
    set_no_list = def_dict.get("Number of defined quantum number sets", None)

    code_list = def_dict.get("A code that defines this set of quantum numbers", None)
    line_no_list = def_dict.get("No. of lines in the broad that contain this code", None)
    qn_no_list = def_dict.get("No. of quantum numbers defined", None)

    qns_list = def_dict.get("Defined quantum number", None)

    if label_list and fname_list and max_J_list and lor_hw_list and temp_exp_list and set_no_list:
        assert len(label_list) == len(fname_list) == len(max_J_list) == len(lor_hw_list) == len(temp_exp_list) == len(set_no_list), "Broadener lists are not of the same length."
        broadeners = []
        for i in range(len(label_list)):
            dict = {
                "Label for a particular broadener": label_list[i],
                "Filename of particular broadener": fname_list[i],
                "Maximum J for which pressure broadening parameters provided": max_J_list[i],
                'Value of Lorentzian half-width for J" > Jmax': lor_hw_list[i],
                'Value of temperature exponent for lines with J" > Jmax': temp_exp_list[i],
                "Number of defined quantum number sets": set_no_list[i],
                "Quantum number sets": []
            }
            if code_list and line_no_list and qn_no_list:
                length = np.sum(np.array([int(x) for x in set_no_list]))
                assert len(code_list) == len(line_no_list) == len(qn_no_list) == length, "Broadener code lists are not of the same length."
                # Calculate the offset for this broadener
                set_offset = sum(int(set_no_list[m]) for m in range(i))
                for j in range(int(set_no_list[i])):
                    idx = set_offset + j
                    qn_dict = {
                        "A code that defines this set of quantum numbers": code_list[idx],
                        "No. of lines in the broad that contain this code": line_no_list[idx],
                        "No. of quantum numbers defined": qn_no_list[idx],
                        "Defined quantum number": []
                    }
                    if qns_list:
                        length = np.sum(np.array([int(x) for x in qn_no_list]))
                        assert len(qns_list) == length, "Broadener quantum number lists are not of the same length."
                        # Calculate the correct offset for qns_list
                        qn_offset = sum(int(qn_no_list[m]) for m in range(idx))
                        for k in range(int(qn_no_list[idx])):
                            qn_dict["Defined quantum number"].append(qns_list[qn_offset + k])
                    dict["Quantum number sets"].append(qn_dict)
            broadeners.append(dict)

        def_dict["Broadening parameters"] = broadeners
        def_dict.pop("Label for a particular broadener", None)
        def_dict.pop("Filename of particular broadener", None)
        def_dict.pop("Filename for a particular broadener", None)
        def_dict.pop("Maximum J for which pressure broadening parameters provided", None)
        def_dict.pop("Value of Lorentzian half-width for J\" > Jmax", None)
        def_dict.pop("Value of temperature exponent for lines with J\" > Jmax", None)
        def_dict.pop("Number of defined quantum number sets", None)
        def_dict.pop("A code that defines this set of quantum numbers", None)
        def_dict.pop("No. of lines in the broad that contain this code", None)
        def_dict.pop("No. of quantum numbers defined", None)
        def_dict.pop("Defined quantum number", None)

    def_dict["Quantum labels"] = quantum_labels
    def_dict["Auxiliary labels"] = auxiliary_labels
    return def_dict

# Output formatting
def line_formatter(value, desc):
    """Formats a line for the new definition file."""
    value = str(value).strip()
    desc = str(desc).strip()
    if value == "None" or value == None:
        if desc == "Quantum case label":
            return ""
        else:
            error_log(f"Dataset: {filename} || Value for '{desc}' is None.", "Warn")
            return ""
    l = "EXOMOL.def                                                                      # ID"
    try:
        tot_chars = int(len(l.split("#")[0]))
        num_chars = int(len(value))
        padding = tot_chars - num_chars
        assert padding >= 0, "Padding cannot be negative."
    except Exception as e:
        error_log(f"Dataset: {filename} || {e} for {desc}", "Warn")
        padding = 0
    line = value + " " * padding + "# " + desc + "\n"
    return line

def label_formatter(labels_dict):
    """Formats the quantum labels for the definition file."""
    result = ""
    try:
        for i, dict in enumerate(labels_dict):
            keys = list(dict.keys())
            for key in keys:
                result += line_formatter(dict[key], f"{key} {i+1}")
    except Exception as e:
        error_log(f"Dataset: {filename} || {e}", "Critical")
        result = ""
    return result

def broadener_formatter(broad_dict):
    """Formats the broadening parameters for the definition file."""
    result = ""
    try:
        for dict in broad_dict:
            keys = list(dict.keys())
            for key in keys:
                if isinstance(dict[key], list):
                    for value in dict[key]:
                        sub_keys = list(value.keys())
                        for sub_key in sub_keys:
                            if isinstance(value[sub_key], list):
                                for sub_value in value[sub_key]:
                                    result += line_formatter(sub_value, f"{sub_key}")
                            else:
                                result += line_formatter(value[sub_key], f"{sub_key}")
                else:
                    result += line_formatter(dict[key], f"{key}")
    except Exception as e:
        error_log(f"Dataset: {filename} || {e}", "Critical")
        result = ""
    return result
    
def bool_formmatter(bools_dict):
    """Formats the boolean flags for the definition file."""
    result = ""
    try:
        for key, value in bools_dict.items():
            result += line_formatter(value, f"{key} (1=yes, 0=no)")
    except Exception as e:
        error_log(f"Dataset: {filename} || {e}", "Critical")
        result = ""
    return result

# Update def_dict according to states_labels.json
def def_dict_update(mol, def_dict, labels_info):
    """Updates the definition dictionary with new labels."""
    labels_list = []
    for dict in labels_info:
        if isinstance(dict, str):
            labels_list.append(dict)
        else:
            labels_list.append(dict["Quantum label"])
    backend_master = load_backend_master_file()

    bools = {
        "Lifetime availability": 0,
        "Lande g-factor availability": 0,
        "Uncertainty availability": 0,
        "Hyperfine resolved dataset": 0
    }

    keys = list(def_dict.keys())
    for key in keys:
        if "Lifetime availability" in key:
            def_dict.pop(key, None)
        if "Lande g-factor availability" in key:
            def_dict.pop(key, None)
        if "Uncertainty availability" in key:
            def_dict.pop(key, None)
        if "Hyperfine resolved dataset" in key:
            def_dict.pop(key, None)
        if "Inchi key of molecule" in key:
            try:
                iso_formula = def_dict["IsoFormula"]
                inchi = backend_master[mol][iso_formula]["inchi"]
                inchikey = backend_master[mol][iso_formula]["inchikey"]
                def_dict["In-ChI of molecule"] = inchi
                def_dict["In-ChI key of molecule"] = inchikey
                def_dict.pop(key, None)
            except KeyError as e:
                try:
                    inchi = def_dict.get("In-ChI of molecule", None)
                    inchikey = def_dict.get("In-ChI key of molecule", None)
                    def_dict.pop(key, None)
                except KeyError as e:
                    error_log(f"Dataset: {filename} || In-ChI/In-ChI key not found in backend master file: {e}", "Warn")
                    pass

    standard_labels = load_standard_labels()
    if labels_list[3] == "F":
        bools['Hyperfine resolved dataset'] = 1
    labels = labels_list[4:]
    if "tau" in labels:
        bools["Lifetime availability"] = 1
        labels.remove("tau")
    if "gfactor" in labels:
        bools["Lande g-factor availability"] = 1
        labels.remove("gfactor")
    if "unc" in labels:
        bools["Uncertainty availability"] = 1
        labels.remove("unc")
    def_dict["bools"] = bools
    
    aux_labels = [label for label in labels if "Auxiliary" in label]
    if aux_labels:
        def_dict["Auxiliary labels"] = []
        for aux_label in aux_labels:
            label = aux_label.split(":")[1]
            if label == "SourceType":
                dict = {
                    "Auxiliary title": label,
                    "Format title": "A2 %2s",
                    "Description title": "Ma=MARVEL,Ca=Calculated,EH=Effective Hamiltonian,IE=Isotopologue extrapolation"
                }
                def_dict["Auxiliary labels"].append(dict)
            elif label == "Ecal":
                dict = {
                    "Auxiliary title": label,
                    "Format title": "F12.6 %12.6f",
                    "Description title": "Calculated energy in cm-1"
                }
                def_dict["Auxiliary labels"].append(dict)
    new_labels = []
    prefixes = []
    for label in labels:
        if label not in aux_labels and ":" in label:
            prefixes.append(label.split(":")[0])
            new_labels.append(label.split(":")[1])
        elif label not in aux_labels:
            new_labels.append(label)

    old_labels = def_dict.get("Quantum labels", None)
    if not old_labels:
        error_log(f"Dataset: {filename} || No quantum labels found in definition file.", "Critical")
        exit_script()
    fname = def_dict["Iso-slug"] + "__" + def_dict["Isotopologue dataset name"] + ".states"
    states_file_path = os.path.join(".", "input", mol, fname)

    # Build lookup dicts for old and standard labels for O(1) access
    old_label_map = {d["Quantum label"]: d for d in old_labels}
    standard_label_map = {d["Quantum label"]: d for d in standard_labels}

    for i, label in enumerate(new_labels):
        correction_dict = load_correction_dict()
        if label in list(correction_dict.keys()):
            label = correction_dict[label]
        if label == "J":
            J_cfmt, J_ffmt = check_J_format(states_file_path)
            fmt = " ".join([J_ffmt, J_cfmt])
            new_labels[i] = {
                "Quantum label": label,
                "Format quantum label": fmt,
                "Description quantum label": "Total rotational quantum number"
            }
            continue

        # Try old labels, then standard labels, else fallback
        if label in old_label_map:
            new_labels[i] = old_label_map[label]
        elif label in standard_label_map:
            new_labels[i] = standard_label_map[label]
        elif isinstance(new_labels[i], str):
            new_labels[i] = {
                "Quantum label": label,
                "Format quantum label": detect_format(states_file_path, i),
                "Description quantum label": ""
            }
            for dict in labels_info:
                if not isinstance(dict, str):
                    if dict["Quantum label"] == label:
                        new_labels[i]["Description quantum label"] = dict["Description quantum label"]
                        break

            error_log(f"Dataset: {filename} || Quantum label '{label}' not found in old labels. Please input the description manually.")


    def_dict["Quantum labels"] = new_labels
    def_dict["No. of quanta defined"] = len(new_labels)

def update_def(def_file_path, def_dict):
    """Creates a new definition file with updated labels."""
    os.makedirs(os.path.join(os.getcwd(), "output"), exist_ok=True)
    mol = os.path.basename(os.path.dirname(def_file_path))
    output_dir = os.path.join(os.getcwd(), "output", mol)
    os.makedirs(output_dir, exist_ok=True)
    output_file_path = os.path.join(output_dir, os.path.basename(def_file_path))
    
    keys = list(def_dict.keys())

    with open(def_file_path, 'r') as def_file:
        with open(output_file_path, 'w') as output_file:
            lines = def_file.readlines()
            for l in lines:
                if "#" not in l:
                    continue
                parts = [entry.strip() for entry in l.split("#", 1)]
                if len(parts) != 2:
                    continue
                value, desc = parts
                if "In-ChI" in desc:
                    continue
                if desc in keys and not isinstance(def_dict.get(desc), list):
                    value = def_dict.get(desc)
                    output_file.write(line_formatter(value, desc))
                    def_dict.pop(desc, None)
                elif desc in keys and isinstance(def_dict.get(desc), list):
                    for value in def_dict[desc]:
                        output_file.write(line_formatter(value, desc))
                    def_dict.pop(desc, None)         

                if desc == "No. of quanta defined":
                    output_file.write(label_formatter(def_dict.get("Quantum labels", [])))
                    output_file.write(label_formatter(def_dict.get("Auxiliary labels", [])))
                    def_dict.pop("Quantum labels", None)
                    def_dict.pop("Auxiliary labels", None)
                elif desc == "Number of irreducible representations":
                    output_file.write(label_formatter(def_dict.get("irreps", [])))
                    def_dict.pop("irreps", None)
                elif desc == "Default value of temperature exponent for all lines":
                    output_file.write(broadener_formatter(def_dict.get("Broadening parameters", [])))
                    def_dict.pop("Broadening parameters", None)
                elif desc == "Number of atoms":
                    output_file.write(label_formatter(def_dict.get("Isotope information", [])))
                    def_dict.pop("Isotope information", None)
                elif desc == "No. of quanta cases":
                    case_labels = def_dict.get("Quantum case label", [])
                    if isinstance(case_labels, list):
                        for i, case in enumerate(case_labels):
                            output_file.write(line_formatter(case, f"Quantum case label {i+1}"))
                    else:
                        output_file.write(line_formatter(case_labels, f"Quantum case label"))
                    def_dict.pop("Quantum case label", None)
                elif desc == "Quantum case label":
                    continue
                elif desc == "No. of k-coefficient files available":
                    output_file.write(bool_formmatter(def_dict.get("bools", {})))
                    def_dict.pop("bools", None)
                elif desc == "Version number with format YYYYMMDD":
                    output_file.write(line_formatter(def_dict.get("In-ChI of molecule"), "In-ChI of molecule"))
                    output_file.write(line_formatter(def_dict.get("In-ChI key of molecule"), "In-ChI key of molecule"))
                    def_dict.pop("In-ChI of molecule", None)
                    def_dict.pop("In-ChI key of molecule", None)

                
            # Write any remaining entries in def_dict to the output file
            for desc, value in def_dict.items():
                if isinstance(value, list):
                    for v in value:
                        output_file.write(line_formatter(v, desc))
                else:
                    output_file.write(line_formatter(value, desc))
                    print(value, desc)

def make_label_json():
    """Creates a JSON file from the label definitions in the Excel file."""
    excel_path = os.path.join(os.getcwd(), "label_editor.xlsx")
    try:
        wb = openpyxl.load_workbook(excel_path, data_only=True)
    except Exception as e:
        error_log(f"Failed to load Excel file '{excel_path}': {e}", "Critical")
        return

    # Load def_states_check.csv once for all sheets
    csv_path = os.path.join(os.getcwd(), "other_materials", "lib", "def_states_check.csv")
    try:
        with open(csv_path, 'r') as file:
            lines = file.read().splitlines()
        csv_data = [line.split(",") for line in lines[::4]]
    except Exception as e:
        error_log(f"Failed to load CSV file '{csv_path}': {e}", "Critical")
        csv_data = []

    all_label_dicts = []

    for ws in tqdm(wb.worksheets, desc="Reading labels from Excel spreadsheet"):
        if ws is None:
            error_log(f"Failed to load worksheet from {excel_path}", "Critical")
            continue

        first_line = ws.cell(row=1, column=1).value
        mol, fname = str(first_line).split(",")
        mol = mol.split(":")[1].strip()
        fname = fname.split(":")[1].strip()
        iso_slug = fname.split("__")[0]
        dsname = fname.split("__")[1]
        dsname = dsname.replace(".def", "")

        # Find DOI from CSV data
        doi = ""
        for info in csv_data:
            if info[0].strip() == mol and info[1].strip() == iso_slug and info[2].strip() == dsname:
                doi = info[3].strip()
                break

        # Read rows 2-3 as a DataFrame (headers and types)
        try:
            data_rows = [list(row) for row in ws.iter_rows(min_row=2, max_row=3, values_only=True)]
            data_array = np.array(data_rows, dtype=object).T
        except Exception as e:
            error_log(f"Failed to create data array from rows 2-3 in sheet {ws.title}: {e}", "Critical")
            continue

        # Read descriptions from row 7 onwards
        try:
            desc_rows = [list(row) for row in ws.iter_rows(min_row=7, values_only=True)]
            desc_dict = {
                str(row[0]).strip(): str(row[1]).strip() if row[1] is not None else ""
                for row in desc_rows if row[0] is not None
            }
        except Exception as e:
            error_log(f"Failed to create description dict from rows 7+ in sheet {ws.title}: {e}", "Critical")
            desc_dict = {}

        label_dict = {
            "mol": mol,
            "iso_slug": iso_slug,
            "doi": doi,
            "linelist": dsname,
            "ds_name": f"{iso_slug}__{dsname}",
            "labels": []
        }
        for i in range(data_array.shape[0]):
            label_name = str(data_array[i, 0]).strip()
            label = {
                "Quantum label": label_name,
                "Format quantum label": str(data_array[i, 1]).strip(),
                "Description quantum label": desc_dict.get(label_name, "")
            }
            if label["Description quantum label"] == "":
                error_log(f"Missing description for label '{label_name}' in sheet {ws.title}. Please check if you inputted the description correctly in the Excel file.")
            label_dict["labels"].append(label)
        all_label_dicts.append(label_dict)

    # Save to JSON
    out_path = os.path.join(os.getcwd(), "other_materials", "lib", "labels.json")
    try:
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(all_label_dicts, f, indent=4, ensure_ascii=False)
    except Exception as e:
        error_log(f"Failed to write labels JSON to '{out_path}': {e}", "Critical")

def main():
    global date
    date = datetime.now().strftime("%Y%m%dT%H%M%S")
    error_log("Script started.", "Log")
    def_folder_path = os.path.join(os.getcwd(), "input")
    if not os.path.exists(def_folder_path):
        os.makedirs(def_folder_path)
        print("Input folder has been created. Please place your .def files in the 'input' folder.")
        exit_script()
    make_label_json()
    def_paths = []
    labels_data = load_labels_data()
    for root, _, files in os.walk(def_folder_path):
        for file in files:
            if file.endswith(".def"):
                def_paths.append(os.path.abspath(os.path.join(root, file)))

    if not labels_data:
        print(f"No updated labels file found in the folder. Please ensure 'states_labels.json' is correctly named and present.")
        exit_script()

    if not def_paths:
        print(f"No definition files found in the folder: {def_folder_path}")
        exit_script()
    
    skipped_files = []
    for def_file_path in tqdm(def_paths, desc="Processing definition files"):
        global filename
        filename = os.path.basename(def_file_path)
        filename = filename.replace(".def", "")
        if labels_data is not None:
            match = next((item for item in labels_data if item.get('ds_name') == filename), None)
        else:
            match = None
        if match:
            mol = match['mol']
            f = match['ds_name']
            download_states_first_line(mol, f)
            def_dict = read_def_file(def_file_path)
            labels_list = match['labels']
            def_dict_update(mol, def_dict, labels_list)
            update_def(def_file_path, def_dict)
        else:
            skipped_files.append(filename)
            continue


    if skipped_files:
        print(f"Skipped {len(skipped_files)} files due to missing or incomplete labels:")
        for filename in skipped_files:
            print(f" - {filename}")

if __name__ == "__main__":
    os.chdir(os.path.dirname(os.path.abspath(__file__)))
    log_file_path = os.path.join(os.path.dirname(__file__), "log.txt")
    if os.path.exists(log_file_path):
        os.remove(log_file_path)
    main()
    make_def_json()
    exit_script()
    